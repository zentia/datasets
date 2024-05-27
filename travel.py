from openpyxl import load_workbook
import json

# 指定Excel文件路径
file_path = 'test.xlsx'
json_path = 'travel.json'

# 加载工作簿
workbook = load_workbook(filename=file_path)

sheet_name = '景点表'

# 选择活动的工作表
sheet = workbook[sheet_name]

data = []

name = ''
end = False
# 遍历所有行和列
for row in sheet.iter_rows():
    if end:
        break
    for cell in row:
        if cell.column_letter == 'C':
            name = cell.value
            if name == 'name':
                break
            if name is None:
                end = True
                break
        elif cell.column_letter == 'Q':
            if cell.value is not None:
                data.append({
                    "instruction":"你是一个聊天机器人，能理解并回答问题",
                    "input":f"请介绍一下{name}",
                    "output":cell.value
                })
        elif cell.column_letter == 'O':
            data.append({
                "instruction":"你是一个聊天机器人，能理解并回答问题",
                "input":f"{name}在哪里？",
                "output":f"{name}位于{cell.value}"
            })
        elif cell.column_letter == 'X':
            if cell.value is not None:
                data.append({
                    "instruction":"你是一个聊天机器人，能理解并回答问题",
                    "input":f"{name}什么时候开放？",
                    "output":f"{name}开放时间是：{cell.value}"
                })
        elif cell.column_letter == 'Z':
            if cell.value is not None:
                data.append({
                    "instruction":"你是一个聊天机器人，能理解并回答问题",
                    "input":f"介绍一下{name}的停车场信息",
                    "output":f"{name}：{cell.value}"
                })

# 关闭工作簿
workbook.close()

with open(json_path, 'w') as file:
    json.dump(data, file, ensure_ascii=False, indent=4)